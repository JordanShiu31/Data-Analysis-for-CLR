import pandas as pd
import os
import os.path
import numpy as np
import openpyxl

# constants to covert to size
MB = 2
GB = 3

def find_main_folders(root):
    # next will only get the first results the form of a tuple (durpath,dirnames,filenames) and the [1] will return the dirnames
    headers = next(os.walk(root))[1]
    return [os.path.join(root, x) for x in headers]

def create_list_of_all_files_into_csv(root, folder_counter):
    global file_type_list
    df = pd.DataFrame(columns=["Filename", "Size (MB)", "Size (GB)", "File_type","Location"])
    i = 1
    total_size_mb = 0
    total_size_gb = 0
    # loops in order dirpath -> dirnames[1] -> filenames[1] etc
    for dirpath, dirnames, filenames in os.walk(root):
        # print(filenames)
        for file in filenames:

            file_path = os.path.join(dirpath, file)

            # checks if the path is not a symbolic link
            if not os.path.islink(file_path):
                file_size = os.path.getsize(file_path)
                # converts to  MBs (GB would be 3)
                size_mb = file_size / 1024**MB
                size_gb = file_size / 1024**GB
                _, extension = os.path.splitext(file)
                if extension not in file_type_list:
                    file_type_list.append(extension)
                total_size_mb += size_mb
                total_size_gb += size_gb
            
            # adds files into the dataframe
            df.loc[i, "Filename"] = file
            df.loc[i, "Size (MB)"] = round(size_mb,2)
            df.loc[i, "Size (GB)"] = round(size_gb,2)
            df.loc[i, "File_type"] = extension
            df.loc[i, "Location"] = file_path

            # print(f"Filepath name is: {file} and size is: {si`ze:.2f}")
            i+=1

    # creates a new df that shows only cols size and file type and is grouped by filetype. We aggregate the sum of size and count of file types
    # and we drop the firt row and the reset index to combine to the old df
    # we the change the sum col to float type and also round to 2 dp
    stats_df = df[["Size (MB)", "File_type"]].groupby(["File_type"]).aggregate(["sum","count"]).droplevel(0, axis=1).reset_index()
    stats_df["sum"] = stats_df["sum"].astype(float)
    stats_df["sum"] = np.around(stats_df["sum"], 2)

    # sort values in descending order and reset the row index
    stats_df = stats_df.sort_values(by=["sum"], ascending=False)
    stats_df = stats_df.reset_index(drop=True)

    # reset the index of original df and overwirte it (inplace = True) and then we join with other df
    df.reset_index(inplace=True)
    df = df.join(stats_df, lsuffix="_", rsuffix="_unique")

    # we join before we add the total stats since joining stats_df will always work because stats_df gets the unique fle type which will be always <= total number of files
    # adds the total MB, GB and file count at the end of the df
    if len(stats_df) > 0:
        last_row_index = stats_df.index[-1] + 1
        df.loc[last_row_index, "sum"] = round(total_size_mb,2)
        df.loc[last_row_index + 1, "sum"] = round(total_size_gb,2)
        df.loc[last_row_index, "count"] = stats_df["count"].sum()

    # rename headers
    df.rename(columns={"File_type_unique": "Unique_file_type","sum": "Total size (MB)", "count": "Total number of file type"}, inplace=True)

    # now we clean up the DF by dropping the index named col
    df.drop("index", inplace=True, axis=1)

    # export it to Excel csv
    # df.to_csv(f"C:/Users/jordan.shiu/OneDrive - AECOM/Desktop/File Size App/test_{folder_counter}.csv", index=False)
    return df

# root = "C:/Users/jordan.shiu/OneDrive - AECOM//0. Working Folder - CLR2 TA - Movement Workstream"
# root = "C:/Users/Jordan/OneDrive/Documents/University 2021 Feb update/UNSW 2021 Sem 1"
root = "C:/Users/jordan.shiu/OneDrive - AECOM/Documents"
list_of_main_folders = find_main_folders(root)
folder_counter = 1
excel_file_name = "exceltest_docs.xlsx"
sheet_names = ["0 Data Received", "0_Data Sent", "1_Reports", "2_Benchmark Model", "3_Options Testing", "4_Demand Check", "5_Int Volumes", "6_Model Review", "7_Reporting", 
               "8_VLC Demand", "9_Future BC Arup Model", "10_PC Arup Model", "11_NCA Board", "12_2017 BC Recalibration", "13_2019 CLR Model", "14_Future BC and PC", "15_Final s2a Models", "16_s2b Model", "Consturction Modelling"]
file_type_list = []
list_of_excel_sheet_df_contents = []

# creates the excel if it doesnt exists
if not os.path.exists(excel_file_name):
    print("works")
    # createst the excel and saves it to the relative path
    workbook = openpyxl.Workbook()
    workbook.save(excel_file_name)

# new_workbook = openpyxl.load_workbook("C:/Users/jordan.shiu/OneDrive - AECOM/Documents/Grad 2021/09262023 - CLR Stage 2a_2b/11202023 - Sharepoint Folders/Book1.xlsx")

j = 0
# Excel writer allows df to be pasted into a sheet and is looped to paste all sheets (MAKE SURE length of sheename list matches or > than folder list or it wil break)
with pd.ExcelWriter(excel_file_name) as writer:
    for folder_path in list_of_main_folders:
        if os.path.exists(folder_path):
            curr_df = create_list_of_all_files_into_csv(folder_path, folder_counter)
            list_of_excel_sheet_df_contents.append(curr_df)
            curr_df.to_excel(writer, sheet_name=sheet_names[j], index=False)
            j+=1
            print(j)
            # print(curr_df)
        else:
            print(f"The file: {folder_path} does not exist")
            break
        folder_counter +=1

    # Convert list containing all unique file types and rename the header
    file_type_df = pd.DataFrame(file_type_list)
    file_type_df.rename(columns={0: "Unique_file_type"}, inplace=True)
    index=0
    # loop through all DFs (each excel sheet) and merge the summarised data into the summary DF
    # Rename the Total size and total no of file type headers to something else since we cant merge duplicated headers
    for df in list_of_excel_sheet_df_contents:
        file_type_df = pd.merge(file_type_df, df[["Unique_file_type", "Total size (MB)","Total number of file type"]],  how="left", on="Unique_file_type")
        file_type_df.rename(columns={"Total size (MB)": f"{sheet_names[index]}_size_(MB)","Total number of file type":f"{sheet_names[index]}_no_files"}, inplace=True)
        index+=1
    #now we write the summary df to excel
    file_type_df.to_excel(writer, sheet_name="Summary", index=False)

# Moves the Summary tab to the front of excel
workbook = openpyxl.load_workbook(excel_file_name)
workbook.move_sheet("Summary",-(len(workbook.sheetnames)-1))
workbook.save(excel_file_name)
workbook.close()
